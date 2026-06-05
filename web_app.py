import streamlit as st
import os
import re
import shutil
import requests
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage, ExifTags
from io import BytesIO

# ==========================================
# MODERN DARK MODE DESIGN
# ==========================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

/* ---- GLOBAL ---- */
html, body, [data-testid="stAppViewContainer"], [data-testid="stHeader"] {
    background: #0a0a0f !important;
    color: #e0e0e6 !important;
    font-family: 'Space Grotesk', sans-serif !important;
}

[data-testid="stAppViewContainer"] > .main {
    background: #0a0a0f !important;
}

/* Subtle grain overlay */
[data-testid="stAppViewContainer"]::before {
    content: '';
    position: fixed;
    inset: 0;
    opacity: 0.025;
    pointer-events: none;
    z-index: 9999;
    background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='noise'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23noise)'/%3E%3C/svg%3E");
}

/* ---- SCROLLBAR ---- */
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #0a0a0f; }
::-webkit-scrollbar-thumb { background: #2a2a3a; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #3a3a5a; }

/* ---- TITLE ---- */
[data-testid="stMarkdownContainer"] h1 {
    font-family: 'Space Grotesk', sans-serif !important;
    font-weight: 700 !important;
    font-size: 2.4rem !important;
    background: linear-gradient(135deg, #a78bfa, #7c3aed, #c084fc);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    letter-spacing: -0.02em;
    padding-bottom: 0.3rem !important;
    border-bottom: 2px solid rgba(124, 58, 237, 0.2);
}

/* ---- CAPTION ---- */
[data-testid="stCaption"] {
    color: #6b6b80 !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.78rem !important;
    letter-spacing: 0.04em;
}

/* ---- SUBHEADERS ---- */
[data-testid="stMarkdownContainer"] h3 {
    font-family: 'Space Grotesk', sans-serif !important;
    font-weight: 600 !important;
    font-size: 1.15rem !important;
    color: #c4b5fd !important;
    letter-spacing: 0.01em;
    padding-top: 0.5rem !important;
}

/* ---- DIVIDER ---- */
hr {
    border: none !important;
    height: 1px !important;
    background: linear-gradient(90deg, transparent, #2a2a4a, transparent) !important;
    margin: 1.5rem 0 !important;
}

/* ---- INPUTS ---- */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input {
    background: #131320 !important;
    border: 1px solid #2a2a40 !important;
    border-radius: 10px !important;
    color: #e8e8f0 !important;
    font-family: 'Space Grotesk', sans-serif !important;
    padding: 0.7rem 1rem !important;
    transition: border-color 0.25s ease, box-shadow 0.25s ease !important;
}

[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus {
    border-color: #7c3aed !important;
    box-shadow: 0 0 0 3px rgba(124, 58, 237, 0.15) !important;
    outline: none !important;
}

[data-testid="stTextInput"] input::placeholder {
    color: #4a4a60 !important;
    font-style: italic;
}

/* ---- LABELS ---- */
[data-testid="stWidgetLabel"] label,
label {
    color: #a0a0b8 !important;
    font-family: 'Space Grotesk', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.85rem !important;
    letter-spacing: 0.02em;
}

/* ---- SELECTBOX / DROPDOWN ---- */
[data-testid="stSelectbox"] > div > div {
    background: #131320 !important;
    border: 1px solid #2a2a40 !important;
    border-radius: 10px !important;
    color: #e8e8f0 !important;
    transition: border-color 0.25s ease !important;
}

[data-testid="stSelectbox"] > div > div:hover {
    border-color: #7c3aed !important;
}

/* ---- RADIO BUTTONS ---- */
[data-testid="stRadio"] > div {
    gap: 0.5rem !important;
}

[data-testid="stRadio"] label {
    background: #131320 !important;
    border: 1px solid #2a2a40 !important;
    border-radius: 8px !important;
    padding: 0.45rem 0.9rem !important;
    transition: all 0.2s ease !important;
}

[data-testid="stRadio"] label:hover {
    border-color: #7c3aed !important;
    background: #1a1a2e !important;
}

/* ---- FILE UPLOADER ---- */
[data-testid="stFileUploader"] {
    border: 2px dashed #2a2a4a !important;
    border-radius: 14px !important;
    background: rgba(124, 58, 237, 0.03) !important;
    padding: 1rem !important;
    transition: border-color 0.3s ease, background 0.3s ease !important;
}

[data-testid="stFileUploader"]:hover {
    border-color: #7c3aed !important;
    background: rgba(124, 58, 237, 0.06) !important;
}

[data-testid="stFileUploader"] button {
    background: #1e1e35 !important;
    color: #c4b5fd !important;
    border: 1px solid #3a3a5a !important;
    border-radius: 8px !important;
    font-family: 'Space Grotesk', sans-serif !important;
    font-weight: 500 !important;
    transition: all 0.2s ease !important;
}

[data-testid="stFileUploader"] button:hover {
    background: #7c3aed !important;
    color: #fff !important;
    border-color: #7c3aed !important;
}

/* ---- PRIMARY BUTTONS ---- */
[data-testid="stButton"] button[kind="primary"],
[data-testid="stDownloadButton"] a[kind="primary"],
[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #7c3aed, #6d28d9) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 12px !important;
    font-family: 'Space Grotesk', sans-serif !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    padding: 0.75rem 1.5rem !important;
    letter-spacing: 0.02em;
    transition: all 0.3s ease !important;
    box-shadow: 0 4px 20px rgba(124, 58, 237, 0.3) !important;
}

[data-testid="stButton"] button[kind="primary"]:hover,
[data-testid="stDownloadButton"] a:hover,
[data-testid="stDownloadButton"] button:hover {
    background: linear-gradient(135deg, #8b5cf6, #7c3aed) !important;
    box-shadow: 0 6px 30px rgba(124, 58, 237, 0.5) !important;
    transform: translateY(-1px);
}

/* ---- SECONDARY / DANGER BUTTONS ---- */
[data-testid="stButton"] button:not([kind="primary"]) {
    background: #1a1a2e !important;
    color: #a78bfa !important;
    border: 1px solid #2a2a4a !important;
    border-radius: 10px !important;
    font-family: 'Space Grotesk', sans-serif !important;
    font-weight: 500 !important;
    transition: all 0.2s ease !important;
}

[data-testid="stButton"] button:not([kind="primary"]):hover {
    background: #1e1e38 !important;
    border-color: #7c3aed !important;
    color: #c4b5fd !important;
}

/* ---- SUCCESS / WARNING / ERROR / INFO ---- */
[data-testid="stAlert"] {
    border-radius: 10px !important;
    border: none !important;
    font-family: 'Space Grotesk', sans-serif !important;
    font-size: 0.88rem !important;
}

[data-testid="stAlert"][kind="success"] {
    background: rgba(16, 185, 129, 0.1) !important;
    color: #6ee7b7 !important;
    border-left: 3px solid #10b981 !important;
}

[data-testid="stAlert"][kind="warning"] {
    background: rgba(245, 158, 11, 0.1) !important;
    color: #fcd34d !important;
    border-left: 3px solid #f59e0b !important;
}

[data-testid="stAlert"][kind="error"] {
    background: rgba(239, 68, 68, 0.1) !important;
    color: #fca5a5 !important;
    border-left: 3px solid #ef4444 !important;
}

[data-testid="stAlert"][kind="info"] {
    background: rgba(99, 102, 241, 0.08) !important;
    color: #a5b4fc !important;
    border-left: 3px solid #6366f1 !important;
}

/* ---- TOAST ---- */
[data-testid="stToast"] {
    background: #1a1a2e !important;
    color: #a78bfa !important;
    border: 1px solid #2a2a4a !important;
    border-radius: 10px !important;
    font-family: 'Space Grotesk', sans-serif !important;
}

/* ---- SPINNER ---- */
[data-testid="stSpinner"] {
    color: #a78bfa !important;
}

/* ---- COLUMNS GAP ---- */
[data-testid="stHorizontalBlock"] {
    gap: 1.2rem !important;
}

/* ---- SIDEBAR (hide or style) ---- */
[data-testid="stSidebar"] {
    background: #0d0d18 !important;
    border-right: 1px solid #1a1a2e !important;
}

/* ---- NUMBER INPUT STEPPERS ---- */
[data-testid="stNumberInput"] button {
    background: #1e1e35 !important;
    color: #a78bfa !important;
    border: 1px solid #2a2a4a !important;
}

[data-testid="stNumberInput"] button:hover {
    background: #7c3aed !important;
    color: #fff !important;
}

/* ---- CONTAINER PADDING ---- */
.block-container {
    padding-top: 2rem !important;
    padding-bottom: 2rem !important;
    max-width: 780px !important;
}

/* ---- ANIMATED ENTRANCE ---- */
[data-testid="stVerticalBlock"] > div {
    animation: fadeSlideUp 0.5s ease forwards;
    opacity: 0;
}

@keyframes fadeSlideUp {
    from { opacity: 0; transform: translateY(16px); }
    to   { opacity: 1; transform: translateY(0); }
}

[data-testid="stVerticalBlock"] > div:nth-child(1) { animation-delay: 0.0s; }
[data-testid="stVerticalBlock"] > div:nth-child(2) { animation-delay: 0.05s; }
[data-testid="stVerticalBlock"] > div:nth-child(3) { animation-delay: 0.10s; }
[data-testid="stVerticalBlock"] > div:nth-child(4) { animation-delay: 0.15s; }
[data-testid="stVerticalBlock"] > div:nth-child(5) { animation-delay: 0.20s; }
[data-testid="stVerticalBlock"] > div:nth-child(6) { animation-delay: 0.25s; }
[data-testid="stVerticalBlock"] > div:nth-child(7) { animation-delay: 0.30s; }
[data-testid="stVerticalBlock"] > div:nth-child(8) { animation-delay: 0.35s; }
[data-testid="stVerticalBlock"] > div:nth-child(9) { animation-delay: 0.40s; }
[data-testid="stVerticalBlock"] > div:nth-child(10) { animation-delay: 0.45s; }
[data-testid="stVerticalBlock"] > div:nth-child(11) { animation-delay: 0.50s; }
[data-testid="stVerticalBlock"] > div:nth-child(12) { animation-delay: 0.55s; }
[data-testid="stVerticalBlock"] > div:nth-child(13) { animation-delay: 0.60s; }
[data-testid="stVerticalBlock"] > div:nth-child(14) { animation-delay: 0.65s; }
[data-testid="stVerticalBlock"] > div:nth-child(15) { animation-delay: 0.70s; }

/* ---- GLOW ACCENT ON PRIMARY DOWNLOAD ---- */
[data-testid="stDownloadButton"] a,
[data-testid="stDownloadButton"] button {
    animation: pulseGlow 2s ease-in-out infinite;
}

@keyframes pulseGlow {
    0%, 100% { box-shadow: 0 4px 20px rgba(124, 58, 237, 0.3); }
    50%      { box-shadow: 0 4px 30px rgba(124, 58, 237, 0.6); }
}

/* ---- DATAFRAME / TABLE ---- */
[data-testid="stDataFrame"] {
    border-radius: 10px;
    overflow: hidden;
}
</style>
""", unsafe_allow_html=True)

# ==========================================
# KONFIGURASI HALAMAN WEB
# ==========================================
st.set_page_config(page_title="QC Master Pro - LGJA", page_icon="🚀", layout="centered")

def log_traffic(user, toko, tgl_qc, layout):
    JSONBLOB_ID = "019e8740-72c4-7731-8328-0e2c67465233"
    API_URL = f"https://jsonblob.com/api/jsonBlob/{JSONBLOB_ID}"

    try:
        response = requests.get(API_URL)
        try:
            data = response.json()
        except:
            data = []

        if not isinstance(data, list):
            data = []

        tz_jkt = timezone(timedelta(hours=7))
        timestamp = datetime.now(tz_jkt).strftime("%d %B %y / %H:%M")

        data.append({
            "Nama Pengguna": user,
            "Nama Toko": toko,
            "Tanggal QC": tgl_qc,
            "Timestamp": timestamp,
            "Ukuran Layout": layout
        })

        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }
        put_response = requests.put(API_URL, json=data, headers=headers)

        if put_response.status_code in [200, 201]:
            st.toast("✅ Traffic berhasil dicatat di server!")
        else:
            st.error(f"⚠️ Gagal mencatat traffic! Status: {put_response.status_code}")

    except Exception as e:
        st.error(f"⚠️ Sistem Traffic Error Koneksi: {e}")

def correct_orientation(img):
    """Memperbaiki rotasi gambar berdasarkan EXIF Orientation"""
    try:
        if hasattr(img, '_getexif') and img._getexif() is not None:
            exif = img._getexif()
            orientation = exif.get(ExifTags.Base.Orientation)

            if orientation == 3:
                img = img.rotate(180, expand=True)
            elif orientation == 6:
                img = img.rotate(270, expand=True)
            elif orientation == 8:
                img = img.rotate(90, expand=True)
    except:
        pass
    return img

def extract_datetime(filename, uploaded_file_obj):
    match = re.search(r'(\d{4}-\d{2}-\d{2}) at (\d{2}\.\d{2}\.\d{2})', filename)
    if match:
        try:
            return datetime.strptime(f"{match.group(1)} {match.group(2).replace('.', ':')}", "%Y-%m-%d %H:%M:%S")
        except:
            pass
    return datetime.min

# ==========================================
# TAMPILAN ANTARMUKA WEB-APP
# ==========================================
st.title("🚀 QC Master Pro - LGJA")
st.caption("Advanced Dynamic Automation System | Build by AI & GJorma")
st.markdown("---")

# Wajib Isi Nama Pengguna
user_name = st.text_input("👤 Nama Pengguna (Wajib diisi untuk memproses)", placeholder="Masukkan nama Anda...")
if not user_name:
    st.warning("⚠️ Silakan isi Nama Pengguna di atas agar sistem bisa digunakan.")
    st.stop()

# 1. Penamaan File Output
st.subheader("🏷️ 1. Informasi Lokasi QC (Untuk Nama File)")
col1, col2 = st.columns(2)
with col1:
    nama_toko = st.text_input("Nama Toko / Area", placeholder="Contoh: Batavia PIK")
with col2:
    tanggal_qc = st.text_input("Tanggal QC", placeholder="Contoh: 10 Mar 26")

st.markdown("---")

# 2. Upload Template Master
st.subheader("📊 2. Upload Excel Template Master")
mode_template = st.radio("Pilih sumber template Excel:", ["Upload Manual", "File Preset (GitHub)"])

excel_file = None
if mode_template == "Upload Manual":
    excel_file = st.file_uploader("Pilih file Excel (.xlsx)", type=["xlsx"])
else:
    preset_pilihan = st.selectbox("Pilih file preset:",
                                  ["QC_LGJA.xlsx", "QC_Sultan.xlsx", "QC_Vano.xlsx"])
    if os.path.exists(preset_pilihan):
        with open(preset_pilihan, "rb") as f:
            excel_file = BytesIO(f.read())
    else:
        st.error(f"⚠️ File '{preset_pilihan}' belum tersedia di folder server/GitHub Anda.")

selected_sheet = None
if excel_file:
    try:
        excel_file.seek(0)
        wb_scan = load_workbook(excel_file, read_only=True)
        sheet_names = wb_scan.sheetnames
        wb_scan.close()
        selected_sheet = st.selectbox("🔍 Target Sheet:", sheet_names)
    except Exception as e:
        st.error(f"Gagal membaca struktur Excel: {e}")

# Pemilihan Layout & Ukuran Gambar
st.markdown("##### ⚙️ Pengaturan Layout & Ukuran Gambar")
layout_option = st.selectbox("Pilih Opsi Layout:", ["LGJA", "Sultan", "Vano", "Custom"])

if layout_option == "LGJA":
    ROWS = [2, 4, 6, 8, 10, 12]
    COLS = list(range(1, 13))
    COL_W = 41
    ROW_H = 246
    IMAGE_WIDTH_CM = 6.4
    IMAGE_HEIGHT_CM = 8.30
elif layout_option in ["Sultan", "Vano"]:
    ROWS = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30]
    COLS = [1, 2, 3, 4, 5, 6]
    COL_W = 20.43
    ROW_H = 123.75
    IMAGE_WIDTH_CM = 3.2
    IMAGE_HEIGHT_CM = 4.10
else: # Custom
    c_row, c_col = st.columns(2)
    r_in = c_row.text_input("Rows (pisahkan dengan koma)", "2,4,6,8,10,12")
    c_in = c_col.text_input("Columns (angka 1=A, 2=B, pisahkan dengan koma)", "1,2,3,4,5,6")

    try:
        ROWS = [int(x.strip()) for x in r_in.split(",")]
        COLS = [int(x.strip()) for x in c_in.split(",")]
    except:
        st.error("Format Rows/Columns salah! Gunakan angka dipisah koma.")
        st.stop()

    c_w, c_h = st.columns(2)
    COL_W = c_w.number_input("Col Width", value=20.43)
    ROW_H = c_h.number_input("Row Height", value=123.75)

    i_w, i_h = st.columns(2)
    IMAGE_WIDTH_CM = i_w.number_input("Image Width (cm)", value=3.2)
    IMAGE_HEIGHT_CM = i_h.number_input("Image Height (cm)", value=4.10)

st.markdown("---")

# 3. Upload Foto-Foto QC
st.subheader("📸 3. Upload Foto QC Lapangan")
st.info("💡 Tekan Tahan / Pilih Banyak Foto Sekaligus dari Galeri HP Anda.")

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

if st.button("🗑️ Delete All Foto (Reset Uploader)"):
    st.session_state.uploader_key += 1
    st.rerun()

uploaded_photos = st.file_uploader("Pilih semua foto sekaligus dari Galeri",
                                   type=["jpg", "jpeg", "png", "webp"],
                                   accept_multiple_files=True,
                                   key=f"photos_{st.session_state.uploader_key}")

if uploaded_photos:
    st.success(f"**Photo : {len(uploaded_photos)}** (total foto dipilih)")
else:
    st.caption("Belum ada foto yang dipilih")

st.markdown("---")

# 4. Tombol Eksekusi
st.subheader("🚀 4. Eksekusi Proses Data")
if st.button("MULAI EXPORT DAN PROSES DATA", type="primary", use_container_width=True):

    if not nama_toko or not tanggal_qc:
        st.warning("⚠️ Silakan isi Nama Toko dan Tanggal QC terlebih dahulu agar nama file rapi!")
    elif not excel_file:
        st.warning("⚠️ Silakan upload atau pilih Excel Template Master terlebih dahulu!")
    elif not uploaded_photos:
        st.warning("⚠️ Silakan pilih foto QC yang akan diproses!")
    elif not selected_sheet:
        st.error("⚠️ Target sheet tidak valid.")
    else:
        with st.spinner("⏳ Sedang memproses dan mengompres foto... Mohon tunggu..."):
            try:
                log_traffic(user_name, nama_toko, tanggal_qc, layout_option)

                excel_file.seek(0)
                wb = load_workbook(excel_file)
                ws = wb[selected_sheet]

                for c in COLS:
                    ws.column_dimensions[chr(64 + c)].width = COL_W
                for r in ROWS:
                    ws.row_dimensions[r].height = ROW_H

                sorted_photos = sorted(
                    uploaded_photos,
                    key=lambda x: extract_datetime(x.name, x),
                    reverse=True
                )

                all_cells = [f"{chr(64 + col)}{row}" for row in ROWS for col in COLS]
                success_count = 0

                temp_dir = "temp_web_photos"
                os.makedirs(temp_dir, exist_ok=True)

                for i in range(min(len(sorted_photos), len(all_cells))):
                    photo = sorted_photos[i]
                    temp_path = os.path.join(temp_dir, f"compressed_img_{i}.jpg")

                    with PILImage.open(photo) as img_pil:
                        img_pil = correct_orientation(img_pil)

                        if img_pil.mode in ("RGBA", "P"):
                            img_pil = img_pil.convert("RGB")

                        img_pil.thumbnail((1280, 1280), PILImage.Resampling.LANCZOS)
                        img_pil.save(temp_path, format="JPEG",
                                   quality=82,
                                   optimize=True,
                                   subsampling=0)

                    img_excel = ExcelImage(temp_path)
                    img_excel.width = int(IMAGE_WIDTH_CM * 37.8)
                    img_excel.height = int(IMAGE_HEIGHT_CM * 37.8)

                    ws.add_image(img_excel, all_cells[i])
                    success_count += 1

                output = BytesIO()
                wb.save(output)
                wb.close()
                output.seek(0)

                shutil.rmtree(temp_dir, ignore_errors=True)

                fin
/* Sembunyikan header bawaan Streamlit */
#MainMenu, header, footer { visibility: hidden; }

/* ── APP HEADER ── */
.qc-header {
    background: white;
    border: 0.5px solid rgba(0,0,0,0.08);
    border-radius: 12px;
    padding: 20px 24px;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    gap: 16px;
}
.qc-badge {
    background: #1D9E75;
    color: white;
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 0.08em;
    padding: 4px 10px;
    border-radius: 4px;
    display: inline-block;
    margin-bottom: 6px;
    text-transform: uppercase;
}
.qc-title {
    font-size: 26px;
    font-weight: 700;
    color: #1a1a1a;
    margin: 0;
    letter-spacing: -0.02em;
    line-height: 1.2;
}
.qc-subtitle {
    font-size: 12px;
    color: #888;
    font-family: 'JetBrains Mono', monospace;
    margin-top: 4px;
}
.qc-logo {
    font-size: 36px;
    margin-left: auto;
}

/* ── STEP CARDS ── */
.step-card {
    background: white;
    border: 0.5px solid rgba(0,0,0,0.08);
    border-radius: 12px;
    margin-bottom: 16px;
    overflow: hidden;
}
.step-header {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 14px 18px;
    background: #fafafa;
    border-bottom: 0.5px solid rgba(0,0,0,0.06);
}
.step-num {
    width: 28px;
    height: 28px;
    background: #1D9E75;
    color: white;
    border-radius: 6px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    font-size: 12px;
    font-weight: 700;
    font-family: 'JetBrains Mono', monospace;
    flex-shrink: 0;
}
.step-title {
    font-size: 14px;
    font-weight: 500;
    color: #1a1a1a;
    margin: 0;
}

/* ── FIELD LABELS ── */
.field-label {
    font-size: 11px;
    font-weight: 500;
    color: #888;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    font-family: 'JetBrains Mono', monospace;
    margin-bottom: 4px;
    display: block;
}

/* ── INPUT FIELDS ── */
.stTextInput > label { display: none !important; }
.stTextInput > div > div > input {
    background: #f8f7f4 !important;
    border: 0.5px solid rgba(0,0,0,0.12) !important;
    border-radius: 8px !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 14px !important;
    color: #1a1a1a !important;
    padding: 10px 14px !important;
}
.stTextInput > div > div > input:focus {
    border-color: #1D9E75 !important;
    box-shadow: 0 0 0 2px rgba(29,158,117,0.12) !important;
}

/* ── SELECTBOX ── */
.stSelectbox > label { display: none !important; }
.stSelectbox > div > div {
    background: #f8f7f4 !important;
    border: 0.5px solid rgba(0,0,0,0.12) !important;
    border-radius: 8px !important;
}

/* ── RADIO ── */
.stRadio > label { display: none !important; }
.stRadio > div { gap: 8px !important; }
.stRadio > div > label {
    background: #f8f7f4 !important;
    border: 0.5px solid rgba(0,0,0,0.12) !important;
    border-radius: 8px !important;
    padding: 8px 16px !important;
    font-size: 13px !important;
    font-family: 'Syne', sans-serif !important;
    transition: all 0.15s !important;
    cursor: pointer !important;
}
.stRadio > div > label:hover {
    border-color: #1D9E75 !important;
    background: #E1F5EE !important;
}
[data-testid="stRadio"] [data-checked="true"] > label {
    background: #E1F5EE !important;
    border-color: #1D9E75 !important;
    color: #085041 !important;
    font-weight: 500 !important;
}

/* ── FILE UPLOADER ── */
.stFileUploader > label { display: none !important; }
[data-testid="stFileUploaderDropzone"] {
    background: #f8f7f4 !important;
    border: 1.5px dashed rgba(0,0,0,0.15) !important;
    border-radius: 12px !important;
    transition: all 0.15s !important;
}
[data-testid="stFileUploaderDropzone"]:hover {
    border-color: #1D9E75 !important;
    background: #E1F5EE !important;
}

/* ── BUTTON UTAMA ── */
.stButton > button[kind="primary"] {
    background: #1D9E75 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 15px !important;
    font-weight: 700 !important;
    padding: 14px !important;
    letter-spacing: 0.02em !important;
    transition: all 0.15s !important;
    width: 100% !important;
}
.stButton > button[kind="primary"]:hover {
    background: #0F6E56 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 14px rgba(29,158,117,0.25) !important;
}
.stButton > button[kind="primary"]:active {
    transform: scale(0.99) !important;
}

/* ── BUTTON SEKUNDER ── */
.stButton > button:not([kind="primary"]) {
    background: white !important;
    border: 0.5px solid rgba(0,0,0,0.15) !important;
    border-radius: 8px !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 13px !important;
    color: #555 !important;
    transition: all 0.15s !important;
}
.stButton > button:not([kind="primary"]):hover {
    border-color: rgba(0,0,0,0.25) !important;
    color: #1a1a1a !important;
}

/* ── DOWNLOAD BUTTON ── */
.stDownloadButton > button {
    background: #1D9E75 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 15px !important;
    font-weight: 700 !important;
    width: 100% !important;
    padding: 14px !important;
    transition: all 0.15s !important;
}
.stDownloadButton > button:hover {
    background: #0F6E56 !important;
}

/* ── ALERT & INFO ── */
.stWarning {
    background: #FAEEDA !important;
    border: 0.5px solid #EF9F27 !important;
    border-radius: 8px !important;
    color: #633806 !important;
}
.stInfo {
    background: #E1F5EE !important;
    border: 0.5px solid #5DCAA5 !important;
    border-radius: 8px !important;
    color: #085041 !important;
}
.stSuccess {
    background: #E1F5EE !important;
    border: 0.5px solid #1D9E75 !important;
    border-radius: 8px !important;
    color: #085041 !important;
}
.stError {
    background: #FCEBEB !important;
    border: 0.5px solid #E24B4A !important;
    border-radius: 8px !important;
    color: #501313 !important;
}

/* ── TOAST ── */
[data-testid="stToast"] {
    background: #E1F5EE !important;
    border: 0.5px solid #1D9E75 !important;
    border-radius: 10px !important;
    color: #085041 !important;
    font-family: 'Syne', sans-serif !important;
}

/* ── NUMBER INPUT ── */
.stNumberInput > label { display: none !important; }
.stNumberInput > div > div > input {
    background: #f8f7f4 !important;
    border: 0.5px solid rgba(0,0,0,0.12) !important;
    border-radius: 8px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 14px !important;
}

/* ── DIVIDER ── */
hr {
    border: none !important;
    border-top: 0.5px solid rgba(0,0,0,0.08) !important;
    margin: 20px 0 !important;
}

/* ── CAPTION ── */
.stCaption {
    font-size: 12px !important;
    color: #aaa !important;
    font-family: 'JetBrains Mono', monospace !important;
}

/* ── PHOTO COUNT BADGE ── */
.photo-stat {
    display: inline-flex;
    align-items: center;
    gap: 10px;
    background: white;
    border: 0.5px solid rgba(0,0,0,0.08);
    border-radius: 10px;
    padding: 12px 18px;
    margin-top: 8px;
}
.photo-stat-num {
    font-size: 28px;
    font-weight: 700;
    color: #1D9E75;
    font-family: 'JetBrains Mono', monospace;
    line-height: 1;
}
.photo-stat-label {
    font-size: 12px;
    color: #888;
    font-family: 'JetBrains Mono', monospace;
}

/* Padding konten utama */
.block-container {
    padding-top: 24px !important;
    padding-bottom: 40px !important;
    max-width: 680px !important;
}
</style>
""", unsafe_allow_html=True)


# ==========================================
# FUNGSI UTAMA
# ==========================================
def log_traffic(user, toko, tgl_qc, layout):
    JSONBLOB_ID = "019e8740-72c4-7731-8328-0e2c67465233"
    API_URL = f"https://jsonblob.com/api/jsonBlob/{JSONBLOB_ID}"
    try:
        response = requests.get(API_URL)
        try:
            data = response.json()
        except:
            data = []
        if not isinstance(data, list):
            data = []
        tz_jkt = timezone(timedelta(hours=7))
        timestamp = datetime.now(tz_jkt).strftime("%d %B %y / %H:%M")
        data.append({
            "Nama Pengguna": user,
            "Nama Toko": toko,
            "Tanggal QC": tgl_qc,
            "Timestamp": timestamp,
            "Ukuran Layout": layout
        })
        headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}
        put_response = requests.put(API_URL, json=data, headers=headers)
        if put_response.status_code in [200, 201]:
            st.toast("✅ Traffic berhasil dicatat di server!")
        else:
            st.error(f"⚠️ Gagal mencatat traffic! Status: {put_response.status_code}")
    except Exception as e:
        st.error(f"⚠️ Sistem Traffic Error Koneksi: {e}")


def correct_orientation(img):
    try:
        if hasattr(img, '_getexif') and img._getexif() is not None:
            exif = img._getexif()
            orientation = exif.get(ExifTags.Base.Orientation)
            if orientation == 3:
                img = img.rotate(180, expand=True)
            elif orientation == 6:
                img = img.rotate(270, expand=True)
            elif orientation == 8:
                img = img.rotate(90, expand=True)
    except:
        pass
    return img


def extract_datetime(filename, uploaded_file_obj):
    match = re.search(r'(\d{4}-\d{2}-\d{2}) at (\d{2}\.\d{2}\.\d{2})', filename)
    if match:
        try:
            return datetime.strptime(f"{match.group(1)} {match.group(2).replace('.', ':')}", "%Y-%m-%d %H:%M:%S")
        except:
            pass
    return datetime.min


# ==========================================
# HEADER
# ==========================================
st.markdown("""
<div class="qc-header">
    <div>
        <span class="qc-badge">v2.0 LGJA</span>
        <div class="qc-title">QC Master Pro</div>
        <div class="qc-subtitle">Advanced Dynamic Automation System · Build by AI & GJorma</div>
    </div>
    <div class="qc-logo">🚀</div>
</div>
""", unsafe_allow_html=True)


# ==========================================
# STEP 0 — IDENTITAS PENGGUNA
# ==========================================
st.markdown("""
<div class="step-header" style="background:white;border:0.5px solid rgba(0,0,0,0.08);border-radius:12px 12px 0 0;margin-bottom:0">
    <span class="step-num">👤</span>
    <span class="step-title">Identitas Pengguna</span>
</div>
""", unsafe_allow_html=True)

with st.container():
    st.markdown('<span class="field-label">Nama Pengguna (wajib diisi)</span>', unsafe_allow_html=True)
    user_name = st.text_input("Nama Pengguna", placeholder="Masukkan nama Anda...", label_visibility="collapsed")

if not user_name:
    st.warning("⚠️ Silakan isi Nama Pengguna di atas agar sistem bisa digunakan.")
    st.stop()

st.markdown("---")


# ==========================================
# STEP 1 — INFORMASI LOKASI
# ==========================================
st.markdown("""
<div class="step-header" style="background:white;border:0.5px solid rgba(0,0,0,0.08);border-radius:12px 12px 0 0">
    <span class="step-num">01</span>
    <span class="step-title">Informasi Lokasi QC — Untuk Nama File</span>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown('<span class="field-label">Nama Toko / Area</span>', unsafe_allow_html=True)
    nama_toko = st.text_input("Nama Toko", placeholder="Contoh: Batavia PIK", label_visibility="collapsed")
with col2:
    st.markdown('<span class="field-label">Tanggal QC</span>', unsafe_allow_html=True)
    tanggal_qc = st.text_input("Tanggal QC", placeholder="Contoh: 10 Mar 26", label_visibility="collapsed")

st.markdown("---")


# ==========================================
# STEP 2 — TEMPLATE EXCEL
# ==========================================
st.markdown("""
<div class="step-header" style="background:white;border:0.5px solid rgba(0,0,0,0.08);border-radius:12px 12px 0 0">
    <span class="step-num">02</span>
    <span class="step-title">Template Excel Master</span>
</div>
""", unsafe_allow_html=True)

st.markdown('<span class="field-label">Sumber Template</span>', unsafe_allow_html=True)
mode_template = st.radio("Sumber Template", ["📤  Upload Manual", "📦  File Preset (GitHub)"], label_visibility="collapsed", horizontal=True)

excel_file = None
if "Upload" in mode_template:
    st.markdown('<span class="field-label">Pilih file Excel (.xlsx)</span>', unsafe_allow_html=True)
    excel_file = st.file_uploader("Excel", type=["xlsx"], label_visibility="collapsed")
else:
    st.markdown('<span class="field-label">Pilih file preset</span>', unsafe_allow_html=True)
    preset_pilihan = st.selectbox("Preset", ["QC_LGJA.xlsx", "QC_Sultan.xlsx", "QC_Vano.xlsx"], label_visibility="collapsed")
    if os.path.exists(preset_pilihan):
        with open(preset_pilihan, "rb") as f:
            excel_file = BytesIO(f.read())
    else:
        st.error(f"⚠️ File '{preset_pilihan}' belum tersedia di folder server.")

selected_sheet = None
if excel_file:
    try:
        excel_file.seek(0)
        wb_scan = load_workbook(excel_file, read_only=True)
        sheet_names = wb_scan.sheetnames
        wb_scan.close()
        st.markdown('<span class="field-label">Target Sheet</span>', unsafe_allow_html=True)
        selected_sheet = st.selectbox("Sheet", sheet_names, label_visibility="collapsed")
    except Exception as e:
        st.error(f"Gagal membaca struktur Excel: {e}")

st.markdown("---")


# ==========================================
# STEP 3 — LAYOUT
# ==========================================
st.markdown("""
<div class="step-header" style="background:white;border:0.5px solid rgba(0,0,0,0.08);border-radius:12px 12px 0 0">
    <span class="step-num">03</span>
    <span class="step-title">Pengaturan Layout & Ukuran Gambar</span>
</div>
""", unsafe_allow_html=True)

st.markdown('<span class="field-label">Pilih Opsi Layout</span>', unsafe_allow_html=True)
layout_option = st.selectbox("Layout", ["LGJA", "Sultan", "Vano", "Custom"], label_visibility="collapsed")

if layout_option == "LGJA":
    ROWS = [2, 4, 6, 8, 10, 12]
    COLS = list(range(1, 13))
    COL_W, ROW_H = 41, 246
    IMAGE_WIDTH_CM, IMAGE_HEIGHT_CM = 6.4, 8.53
elif layout_option in ["Sultan", "Vano"]:
    ROWS = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30]
    COLS = [1, 2, 3, 4, 5, 6]
    COL_W, ROW_H = 20.43, 123.75
    IMAGE_WIDTH_CM, IMAGE_HEIGHT_CM = 3.2, 4.32
else:
    col_r, col_c = st.columns(2)
    with col_r:
        st.markdown('<span class="field-label">Rows (pisahkan koma)</span>', unsafe_allow_html=True)
        r_in = st.text_input("Rows", "2,4,6,8,10,12", label_visibility="collapsed")
    with col_c:
        st.markdown('<span class="field-label">Columns (1=A, 2=B, dst)</span>', unsafe_allow_html=True)
        c_in = st.text_input("Cols", "1,2,3,4,5,6", label_visibility="collapsed")
    try:
        ROWS = [int(x.strip()) for x in r_in.split(",")]
        COLS = [int(x.strip()) for x in c_in.split(",")]
    except:
        st.error("Format Rows/Columns salah! Gunakan angka dipisah koma.")
        st.stop()

    col_w, col_h = st.columns(2)
    with col_w:
        st.markdown('<span class="field-label">Col Width</span>', unsafe_allow_html=True)
        COL_W = st.number_input("ColW", value=20.43, label_visibility="collapsed")
    with col_h:
        st.markdown('<span class="field-label">Row Height</span>', unsafe_allow_html=True)
        ROW_H = st.number_input("RowH", value=123.75, label_visibility="collapsed")

    col_iw, col_ih = st.columns(2)
    with col_iw:
        st.markdown('<span class="field-label">Image Width (cm)</span>', unsafe_allow_html=True)
        IMAGE_WIDTH_CM = st.number_input("ImgW", value=3.2, label_visibility="collapsed")
    with col_ih:
        st.markdown('<span class="field-label">Image Height (cm)</span>', unsafe_allow_html=True)
        IMAGE_HEIGHT_CM = st.number_input("ImgH", value=4.32, label_visibility="collapsed")

st.markdown("---")


# ==========================================
# STEP 4 — UPLOAD FOTO
# ==========================================
st.markdown("""
<div class="step-header" style="background:white;border:0.5px solid rgba(0,0,0,0.08);border-radius:12px 12px 0 0">
    <span class="step-num">04</span>
    <span class="step-title">Upload Foto QC Lapangan</span>
</div>
""", unsafe_allow_html=True)

st.info("💡 Tekan Tahan / Pilih Banyak Foto Sekaligus dari Galeri HP Anda.")

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

if st.button("🗑️ Reset — Hapus Semua Foto"):
    st.session_state.uploader_key += 1
    st.rerun()

uploaded_photos = st.file_uploader(
    "Foto QC",
    type=["jpg", "jpeg", "png", "webp"],
    accept_multiple_files=True,
    key=f"photos_{st.session_state.uploader_key}",
    label_visibility="collapsed"
)

if uploaded_photos:
    st.markdown(f"""
    <div class="photo-stat">
        <div>
            <div class="photo-stat-num">{len(uploaded_photos)}</div>
            <div class="photo-stat-label">foto dipilih</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
else:
    st.caption("Belum ada foto yang dipilih")

st.markdown("---")


# ==========================================
# STEP 5 — EKSEKUSI
# ==========================================
st.markdown("""
<div class="step-header" style="background:white;border:0.5px solid rgba(0,0,0,0.08);border-radius:12px 12px 0 0">
    <span class="step-num">🚀</span>
    <span class="step-title">Eksekusi — Proses & Export Data</span>
</div>
""", unsafe_allow_html=True)

if st.button("⚡ MULAI EXPORT DAN PROSES DATA", type="primary", use_container_width=True):
    if not nama_toko or not tanggal_qc:
        st.warning("⚠️ Silakan isi Nama Toko dan Tanggal QC terlebih dahulu.")
    elif not excel_file:
        st.warning("⚠️ Silakan upload atau pilih Excel Template Master terlebih dahulu.")
    elif not uploaded_photos:
        st.warning("⚠️ Silakan pilih foto QC yang akan diproses.")
    elif not selected_sheet:
        st.error("⚠️ Target sheet tidak valid.")
    else:
        with st.spinner("⏳ Sedang memproses dan mengompres foto... Mohon tunggu..."):
            try:
                log_traffic(user_name, nama_toko, tanggal_qc, layout_option)

                excel_file.seek(0)
                wb = load_workbook(excel_file)
                ws = wb[selected_sheet]

                for c in COLS:
                    ws.column_dimensions[chr(64 + c)].width = COL_W
                for r in ROWS:
                    ws.row_dimensions[r].height = ROW_H

                sorted_photos = sorted(
                    uploaded_photos,
                    key=lambda x: extract_datetime(x.name, x),
                    reverse=True
                )

                all_cells = [f"{chr(64 + col)}{row}" for row in ROWS for col in COLS]
                success_count = 0
                temp_dir = "temp_web_photos"
                os.makedirs(temp_dir, exist_ok=True)

                for i in range(min(len(sorted_photos), len(all_cells))):
                    photo = sorted_photos[i]
                    temp_path = os.path.join(temp_dir, f"compressed_img_{i}.jpg")
                    with PILImage.open(photo) as img_pil:
                        img_pil = correct_orientation(img_pil)
                        if img_pil.mode in ("RGBA", "P"):
                            img_pil = img_pil.convert("RGB")
                        img_pil.thumbnail((1280, 1280), PILImage.Resampling.LANCZOS)
                        img_pil.save(temp_path, format="JPEG", quality=82, optimize=True, subsampling=0)

                    img_excel = ExcelImage(temp_path)
                    img_excel.width = int(IMAGE_WIDTH_CM * 37.8)
                    img_excel.height = int(IMAGE_HEIGHT_CM * 37.8)
                    ws.add_image(img_excel, all_cells[i])
                    success_count += 1

                output = BytesIO()
                wb.save(output)
                wb.close()
                output.seek(0)
                shutil.rmtree(temp_dir, ignore_errors=True)

                final_filename = f"{nama_toko.strip()} {tanggal_qc.strip()}.xlsx"
                st.success(f"✅ Berhasil menyusun & mengompres **{success_count} foto**! File siap diunduh.")

                st.download_button(
                    label="📥 DOWNLOAD FILE HASIL",
                    data=output,
                    file_name=final_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )

            except Exception as e:
                st.error(f"❌ Terjadi kesalahan sistem saat memproses Excel: {e}")
