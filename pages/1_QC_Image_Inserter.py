import streamlit as st
import os
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from io import BytesIO
from style import (
    inject_css, inject_sidebar_brand, inject_footer,
    log_traffic, correct_orientation, extract_datetime,
    get_preset_files, convert_date_to_english,
)

st.set_page_config(page_title="QC Image Inserter", page_icon="📸", layout="wide")
inject_css()
inject_sidebar_brand()

# PAGE HEADER
st.markdown(
    """
    <div class="page-head">
        <div class="page-head-icon">📸</div>
        <h2>QC Image Inserter</h2>
        <p>Susun foto QC lapangan ke dalam template Excel secara otomatis</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# USER INPUT
st.markdown(
    '<div class="ilp-user"><div class="ilp-user-icon">👤</div>'
    '<div class="ilp-user-hint">Masukkan nama Anda untuk memulai</div></div>',
    unsafe_allow_html=True,
)
user_name = st.text_input("Nama Pengguna", placeholder="Ketik nama Anda di sini...")
if not user_name:
    st.warning("Silakan isi Nama Pengguna di atas agar sistem bisa digunakan.")
    st.stop()

# STEP INDICATOR
st.markdown(
    """
    <div class="ilp-steps">
        <div class="ilp-step active"><div class="ilp-step-num">1</div><div class="ilp-step-label">Info Lokasi</div></div>
        <div class="ilp-step-line active"></div>
        <div class="ilp-step active"><div class="ilp-step-num">2</div><div class="ilp-step-label">Template</div></div>
        <div class="ilp-step-line active"></div>
        <div class="ilp-step active"><div class="ilp-step-num">3</div><div class="ilp-step-label">Upload Foto</div></div>
        <div class="ilp-step-line"></div>
        <div class="ilp-step"><div class="ilp-step-num">4</div><div class="ilp-step-label">Export</div></div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ROW 1 - INFO LOKASI & TEMPLATE EXCEL
col_info, col_template = st.columns(2, gap="medium")

with col_info:
    st.markdown(
        """
        <div class="card-head">
            <div class="card-head-icon">📍</div>
            <h3>Informasi Lokasi QC</h3>
            <p>Detail lokasi untuk penamaan file output</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    nama_toko = st.text_input("Nama Toko / Area", placeholder="Contoh: Batavia PIK")
    tanggal_qc = st.text_input("Tanggal QC", placeholder="Contoh: 10 Mar 26")

with col_template:
    st.markdown(
        """
        <div class="card-head">
            <div class="card-head-icon">📊</div>
            <h3>Template Excel</h3>
            <p>Upload atau pilih template master</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    mode_template = st.radio("Sumber template:", ["Upload Manual", "File Preset"], horizontal=True)

    excel_file = None
    preset_name = ""
    if mode_template == "Upload Manual":
        excel_file = st.file_uploader("Pilih file Excel (.xlsx)", type=["xlsx"], label_visibility="collapsed")
        preset_name = "Upload Manual"
    else:
        available_presets = get_preset_files()
        if available_presets:
            preset_pilihan = st.selectbox("Pilih file preset:", available_presets)
            preset_name = preset_pilihan
            preset_path = os.path.join("presets", preset_pilihan)
            with open(preset_path, "rb") as f:
                excel_file = BytesIO(f.read())
        else:
            st.warning("Folder 'presets/' kosong.")

    selected_sheet = None
    all_sheet_names = []
    if excel_file:
        try:
            excel_file.seek(0)
            wb_scan = load_workbook(excel_file, read_only=True)
            all_sheet_names = wb_scan.sheetnames
            wb_scan.close()
            selected_sheet = st.selectbox("Target Sheet (Foto):", all_sheet_names)
        except Exception as e:
            st.error(f"Gagal membaca struktur Excel: {e}")

# CELL WRITING MODE
st.markdown("---")
st.markdown("##### Store Name, Audit & Date QC")
cell_mode = st.radio(
    "Pilih mode pengisian cell Excel (B6, B7, D6):",
    ["Manual", "Auto"],
    horizontal=True,
    help="Manual: Anda isi sendiri di Excel setelah download. Auto: Sistem otomatis isi dari data yang sudah dimasukkan.",
)

cell_target_sheet = None
if cell_mode == "Auto":
    if all_sheet_names:
        default_cell_sheet = "DETAIL AUDIT" if "DETAIL AUDIT" in all_sheet_names else all_sheet_names[0]
        cell_target_sheet = st.selectbox(
            "Target Sheet (Cell B6/B7/D6):",
            all_sheet_names,
            index=all_sheet_names.index(default_cell_sheet) if default_cell_sheet in all_sheet_names else 0,
            help="Pilih sheet tempat menulis Store Name, Tanggal, dan Nama Pengguna.",
        )
        st.markdown(
            f"""<div style="background:rgba(229,50,45,0.06); border:1px solid rgba(229,50,45,0.15);
                 border-radius:10px; padding:12px 16px; margin-top:8px;">
                <p style="color:#fca5a5; font-size:0.82rem; margin:0; line-height:1.6;">
                    Cell <b>B6</b> = Nama Toko &nbsp;|&nbsp;
                    Cell <b>B7</b> = Tanggal QC &nbsp;|&nbsp;
                    Cell <b>D6</b> = Nama Pengguna<br>
                    Target: <b>{cell_target_sheet}</b>
                </p>
            </div>""",
            unsafe_allow_html=True,
        )
    else:
        st.warning("Tidak ada sheet yang terdeteksi.")

# ROW 2 - LAYOUT & UPLOAD FOTO
col_layout, col_foto = st.columns(2, gap="medium")

with col_layout:
    st.markdown(
        """
        <div class="card-head">
            <div class="card-head-icon">⚙️</div>
            <h3>Pengaturan Layout</h3>
            <p>Atur ukuran dan posisi gambar di Excel</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    layout_option = st.selectbox("Opsi Layout:", ["LGJA", "Sultan", "Vano", "Custom"])

    if layout_option == "LGJA":
        ROWS = [2, 4, 6, 8, 10, 12]
        COLS = list(range(1, 13))
        COL_W = 41
        ROW_H = 246
        IMAGE_WIDTH_CM = 6.4
        IMAGE_HEIGHT_CM = 8.30
    elif layout_option in ["Sultan", "Vano"]:
        ROWS = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30]
        COLS = [1, 2, 3, 4, 5, 6]
        COL_W = 20.43
        ROW_H = 123.75
        IMAGE_WIDTH_CM = 3.2
        IMAGE_HEIGHT_CM = 4.10
    else:
        c_row, c_col = st.columns(2)
        r_in = c_row.text_input("Rows (koma)", "2,4,6,8,10,12")
        c_in = c_col.text_input("Columns (koma)", "1,2,3,4,5,6")
        try:
            ROWS = [int(x.strip()) for x in r_in.split(",")]
            COLS = [int(x.strip()) for x in c_in.split(",")]
        except Exception:
            st.error("Format Rows/Columns salah!")
            st.stop()
        c_w, c_h = st.columns(2)
        COL_W = c_w.number_input("Col Width", value=20.43)
        ROW_H = c_h.number_input("Row Height", value=123.75)
        i_w, i_h = st.columns(2)
        IMAGE_WIDTH_CM = i_w.number_input("Image Width (cm)", value=3.2)
        IMAGE_HEIGHT_CM = i_h.number_input("Image Height (cm)", value=4.10)

with col_foto:
    st.markdown(
        """
        <div class="upload-head">
            <div class="upload-head-icon">📸</div>
            <h3>Upload Foto QC</h3>
            <p>Pilih semua foto sekaligus dari galeri HP</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """<div style="background:rgba(245,158,11,0.08); border:1px solid rgba(245,158,11,0.2);
             border-radius:10px; padding:10px 14px; margin-bottom:14px;">
            <p style="color:#fcd34d; font-size:0.78rem; margin:0; line-height:1.5;">
                <b>Tips HP:</b> Gunakan Gallery bawaan HP atau File Manager.
                Google Photos sering memutus koneksi. Jika reconnecting,
                tunggu beberapa detik lalu coba lagi.
            </p>
        </div>""",
        unsafe_allow_html=True,
    )

    if "uploader_key" not in st.session_state:
        st.session_state.uploader_key = 0

    if st.button("Hapus Semua Foto (Reset)"):
        st.session_state.uploader_key += 1
        st.rerun()

    uploaded_photos = st.file_uploader(
        "Pilih semua foto sekaligus dari Galeri",
        type=["jpg", "jpeg", "png", "webp"],
        accept_multiple_files=True,
        key=f"photos_{st.session_state.uploader_key}",
        label_visibility="collapsed",
    )

    if uploaded_photos:
        st.success(f"{len(uploaded_photos)} foto dipilih")
    else:
        st.caption("Belum ada foto yang dipilih")

# ACTION
st.markdown("---")
st.markdown(
    """
    <div class="ilp-action-head">
        <h3>Siap untuk memproses?</h3>
        <p>Pastikan semua data sudah benar sebelum mengeksekusi</p>
    </div>
    """,
    unsafe_allow_html=True,
)

if st.button("MULAI EXPORT DAN PROSES DATA", type="primary", use_container_width=True):
    if not nama_toko or not tanggal_qc:
        st.warning("Silakan isi Nama Toko dan Tanggal QC!")
    elif not excel_file:
        st.warning("Silakan upload atau pilih Excel Template Master!")
    elif not uploaded_photos:
        st.warning("Silakan pilih foto QC!")
    elif not selected_sheet:
        st.error("Target sheet foto tidak valid.")
    elif cell_mode == "Auto" and not cell_target_sheet:
        st.error("Target sheet cell (DETAIL AUDIT) tidak valid.")
    else:
        with st.spinner("Sedang memproses dan mengompres foto..."):
            try:
                template_str = f"{preset_name}, {selected_sheet}"
                log_traffic(user_name, nama_toko, tanggal_qc, template_str, layout_option)

                excel_file.seek(0)
                wb = load_workbook(excel_file)

                # AUTO: Write to cells B6, B7, D6 in target sheet
                if cell_mode == "Auto" and cell_target_sheet:
                    if cell_target_sheet in wb.sheetnames:
                        ws_cell = wb[cell_target_sheet]
                        ws_cell["B6"] = nama_toko
                        ws_cell["B7"] = convert_date_to_english(tanggal_qc)
                        ws_cell["D6"] = user_name

                ws = wb[selected_sheet]

                for c in COLS:
                    ws.column_dimensions[chr(64 + c)].width = COL_W
                for r in ROWS:
                    ws.row_dimensions[r].height = ROW_H

                sorted_photos = sorted(
                    uploaded_photos,
                    key=lambda x: extract_datetime(x.name, x),
                    reverse=True,
                )
                all_cells = [f"{chr(64 + col)}{row}" for row in ROWS for col in COLS]
                success_count = 0

                temp_dir = "temp_web_photos"
                os.makedirs(temp_dir, exist_ok=True)

                for i in range(min(len(sorted_photos), len(all_cells))):
                    photo = sorted_photos[i]
                    temp_path = os.path.join(temp_dir, f"compressed_img_{i}.jpg")

                    with PILImage.open(photo) as img_pil:
                        img_pil = correct_orientation(img_pil)
                        if img_pil.mode in ("RGBA", "P"):
                            img_pil = img_pil.convert("RGB")
                        img_pil.thumbnail((1280, 1280), PILImage.Resampling.LANCZOS)
                        img_pil.save(
                            temp_path, format="JPEG", quality=82,
                            optimize=True, subsampling=0,
                        )

                    img_excel = ExcelImage(temp_path)
                    img_excel.width = int(IMAGE_WIDTH_CM * 37.8)
                    img_excel.height = int(IMAGE_HEIGHT_CM * 37.8)
                    ws.add_image(img_excel, all_cells[i])
                    success_count += 1

                output = BytesIO()
                wb.save(output)
                wb.close()
                output.seek(0)
                shutil.rmtree(temp_dir, ignore_errors=True)

                final_filename = f"{nama_toko.strip()} {tanggal_qc.strip()}.xlsx"

                cell_msg = ""
                if cell_mode == "Auto" and cell_target_sheet:
                    cell_msg = f" | Cell B6/B7/D6 terisi otomatis di sheet '{cell_target_sheet}'"
                st.success(f"Berhasil menyusun {success_count} foto!{cell_msg}")

                st.download_button(
                    label="DOWNLOAD FILE EXCEL",
                    data=output,
                    file_name=final_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Terjadi kesalahan: {e}")

inject_footer()
