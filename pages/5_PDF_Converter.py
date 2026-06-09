import streamlit as st
import os
from io import BytesIO
from style import inject_css, inject_sidebar_brand, inject_footer

st.set_page_config(page_title="PDF Converter", page_icon="📄", layout="wide")
inject_css()
inject_sidebar_brand()

st.markdown(
    """
    <div class="page-head">
        <div class="page-head-icon">📄</div>
        <h2>PDF Converter</h2>
        <p>Ubah file Excel (.xlsx) menjadi PDF siap kirim ke klien atau atasan</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# Find Unicode fonts on system
FONT_REGULAR = None
FONT_BOLD = None

for p in [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    "/usr/share/fonts/TTF/DejaVuSans.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans.ttf",
]:
    if os.path.exists(p):
        FONT_REGULAR = p
        break

for p in [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
]:
    if os.path.exists(p):
        FONT_BOLD = p
        break


def sanitize_text(text):
    """Replace Unicode chars that basic fonts cannot render."""
    if text is None:
        return ""
    text = str(text)
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u2026": "...",
        "\u2022": "-",
        "\u00a0": " ",
        "\u00b0": "deg",
        "\u2122": "TM",
        "\u00ae": "(R)",
        "\u00a9": "(C)",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def set_pdf_font(pdf, style, size):
    """Set font with Unicode fallback."""
    if FONT_REGULAR:
        if style == "B" and FONT_BOLD:
            pdf.set_font("QCF", "B", size)
        elif style == "B" and not FONT_BOLD:
            pdf.set_font("QCF", "", size)
        else:
            pdf.set_font("QCF", "", size)
    else:
        pdf.set_font("Helvetica", style, size)


def safe_cell(pdf, w, h, text, **kwargs):
    """Write cell with sanitized text."""
    pdf.cell(w, h, sanitize_text(text), **kwargs)


def extract_sheet_images(ws):
    """Extract all images from worksheet, including floating/over-cell images."""
    images = []
    if not hasattr(ws, "_images"):
        return images

    for idx, img in enumerate(ws._images):
        try:
            raw_data = None

            # Try _data() method
            try:
                raw_data = img._data()
            except Exception:
                pass

            # Fallback: try reading from ref/path
            if raw_data is None:
                try:
                    if hasattr(img, "ref") and img.ref and os.path.exists(img.ref):
                        with open(img.ref, "rb") as f:
                            raw_data = f.read()
                except Exception:
                    pass

            # Fallback: try _img attribute
            if raw_data is None:
                try:
                    if hasattr(img, "_img"):
                        raw_data = img._img
                except Exception:
                    pass

            if raw_data and len(raw_data) > 100:
                images.append(raw_data)

        except Exception:
            pass

    return images


def image_to_jpeg_bytes(raw_bytes):
    """Convert any image bytes to JPEG bytes via PIL."""
    from PIL import Image as PILImage

    pil_img = PILImage.open(BytesIO(raw_bytes))

    # Handle transparency
    if pil_img.mode in ("RGBA", "LA", "PA"):
        bg = PILImage.new("RGB", pil_img.size, (255, 255, 255))
        alpha = pil_img.split()[-1]
        bg.paste(pil_img, mask=alpha)
        pil_img = bg
    elif pil_img.mode == "P":
        pil_img = pil_img.convert("RGBA")
        bg = PILImage.new("RGB", pil_img.size, (255, 255, 255))
        alpha = pil_img.split()[-1]
        bg.paste(pil_img, mask=alpha)
        pil_img = bg
    elif pil_img.mode != "RGB":
        pil_img = pil_img.convert("RGB")

    buf = BytesIO()
    pil_img.save(buf, format="JPEG", quality=90, optimize=True)
    buf.seek(0)
    return buf, pil_img.size


# MAIN UI
uploaded_excel = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx"])

if uploaded_excel:
    try:
        from openpyxl import load_workbook

        uploaded_excel.seek(0)
        wb = load_workbook(uploaded_excel, data_only=True)
        sheet_names = wb.sheetnames
        selected_sheet = st.selectbox("Pilih Sheet untuk konversi:", sheet_names)

        # Count images
        ws_check = wb[selected_sheet]
        img_count = len(ws_check._images) if hasattr(ws_check, "_images") else 0
        wb.close()

        if img_count > 0:
            st.markdown(
                f"""<div style="background:rgba(16,185,129,0.08); border:1px solid rgba(16,185,129,0.2);
                     border-radius:10px; padding:10px 14px; margin:8px 0;">
                    <p style="color:#6ee7b7; font-size:0.82rem; margin:0;">
                        Ditemukan <b>{img_count} gambar</b> di sheet ini
                        (termasuk floating/over-cell images). Gambar akan dimasukkan ke PDF.
                    </p>
                </div>""",
                unsafe_allow_html=True,
            )

        if st.button("KONVERSI KE PDF", type="primary", use_container_width=True):
            with st.spinner("Mengonversi Excel ke PDF..."):
                uploaded_excel.seek(0)
                wb = load_workbook(uploaded_excel, data_only=True)
                ws = wb[selected_sheet]

                # Get table data
                all_data = []
                for row in ws.iter_rows(values_only=True):
                    all_data.append(row)

                # Extract images
                raw_images = extract_sheet_images(ws)
                wb.close()

                if not all_data and not raw_images:
                    st.error("Sheet kosong, tidak ada data maupun gambar.")
                    st.stop()

                try:
                    from fpdf import FPDF
                    from fpdf.enums import XPos, YPos

                    pdf = FPDF(orientation="L", unit="mm", format="A4")
                    pdf.alias_nb_pages()
                    pdf.set_auto_page_break(auto=True, margin=20)

                    # Register Unicode fonts
                    if FONT_REGULAR:
                        pdf.add_font("QCF", "", FONT_REGULAR)
                    if FONT_BOLD:
                        pdf.add_font("QCF", "B", FONT_BOLD)

                    pdf.add_page()

                    # Title
                    set_pdf_font(pdf, "B", 14)
                    pdf.set_text_color(40, 40, 40)
                    safe_cell(pdf, 0, 10, selected_sheet,
                              new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                    pdf.ln(1)

                    # Subtitle
                    set_pdf_font(pdf, "", 8)
                    pdf.set_text_color(120, 120, 120)
                    safe_cell(pdf, 0, 6, f"Source: {uploaded_excel.name}",
                              new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                    pdf.ln(2)

                    # Separator
                    pdf.set_draw_color(229, 50, 45)
                    pdf.set_line_width(0.5)
                    pdf.line(10, pdf.get_y(), pdf.w - 10, pdf.get_y())
                    pdf.ln(6)

                    # TABLE DATA
                    if all_data:
                        num_cols = max(len(row) for row in all_data)
                        usable_width = pdf.w - 20
                        col_width = min(usable_width / num_cols, 55)
                        row_height = 8

                        def draw_table_header():
                            set_pdf_font(pdf, "B", 8)
                            pdf.set_fill_color(240, 240, 240)
                            pdf.set_text_color(40, 40, 40)
                            pdf.set_draw_color(200, 200, 200)
                            for cell_val in header_row:
                                text = sanitize_text(cell_val)
                                if len(text) > 45:
                                    text = text[:42] + "..."
                                pdf.cell(col_width, row_height, text, border=1, fill=True,
                                         new_x=XPos.RIGHT, new_y=YPos.TOP)
                            pdf.ln()

                        header_row = all_data[0]
                        draw_table_header()

                        set_pdf_font(pdf, "", 7)
                        pdf.set_fill_color(255, 255, 255)
                        pdf.set_text_color(50, 50, 50)

                        for row_data in all_data[1:]:
                            if pdf.get_y() + row_height > pdf.h - 20:
                                pdf.add_page()
                                draw_table_header()
                                set_pdf_font(pdf, "", 7)
                                pdf.set_fill_color(255, 255, 255)
                                pdf.set_text_color(50, 50, 50)

                            for idx, cell_val in enumerate(row_data):
                                if idx >= num_cols:
                                    break
                                text = sanitize_text(cell_val)
                                if len(text) > 45:
                                    text = text[:42] + "..."
                                pdf.cell(col_width, row_height, text, border=1,
                                         new_x=XPos.RIGHT, new_y=YPos.TOP)
                            pdf.ln()

                    # IMAGES
                    if raw_images:
                        pdf.add_page()
                        set_pdf_font(pdf, "B", 12)
                        pdf.set_text_color(40, 40, 40)
                        safe_cell(pdf, 0, 10, f"Embedded Images ({len(raw_images)} total)",
                                  new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        pdf.ln(6)

                        max_w = pdf.w - 20
                        max_h = pdf.h - 40

                        for idx, raw_data in enumerate(raw_images):
                            try:
                                jpeg_buf, (img_w, img_h) = image_to_jpeg_bytes(raw_data)

                                ratio = min(max_w / img_w, max_h / img_h, 1.0)
                                w = img_w * ratio
                                h = img_h * ratio

                                if pdf.get_y() + h > pdf.h - 20:
                                    pdf.add_page()

                                pdf.image(jpeg_buf, x=10, y=pdf.get_y(), w=w, h=h)
                                pdf.ln(h + 4)

                                set_pdf_font(pdf, "", 7)
                                pdf.set_text_color(150, 150, 150)
                                safe_cell(pdf, 0, 5, f"Image {idx + 1} of {len(raw_images)}",
                                          new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                                pdf.ln(6)

                            except Exception as img_err:
                                set_pdf_font(pdf, "", 8)
                                pdf.set_text_color(200, 60, 60)
                                safe_cell(pdf, 0, 8, f"[Image {idx + 1} gagal: {str(img_err)[:60]}]",
                                          new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                                pdf.ln(4)

                    # Page numbers
                    total_pages = pdf.page
                    for pg in range(1, total_pages + 1):
                        pdf.page = pg
                        pdf.set_y(-15)
                        set_pdf_font(pdf, "", 7)
                        pdf.set_text_color(150, 150, 150)
                        pdf.cell(0, 10, f"Page {pg}/{total_pages}", align="C",
                                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

                    pdf_output = bytes(pdf.output())

                    img_msg = f" + {len(raw_images)} gambar" if raw_images else ""
                    st.success(f"Konversi berhasil! {len(all_data)} baris data{img_msg}.")

                    st.download_button(
                        label="DOWNLOAD PDF",
                        data=pdf_output,
                        file_name=f"{selected_sheet}.pdf",
                        mime="application/pdf",
                        type="primary",
                        use_container_width=True,
                    )

                except ImportError:
                    st.error("Library fpdf2 belum terinstall. Pastikan 'fpdf2' ada di requirements.txt.")

    except Exception as e:
        st.error(f"Gagal membaca file Excel: {e}")
else:
    st.caption("Upload file Excel di atas untuk mulai konversi.")

inject_footer()
